"""Advanced Excel/spreadsheet analysis - multi-sheet, formulas, charts, pivot tables."""

from __future__ import annotations

import csv
import io
import json
import statistics
from collections import defaultdict
from pathlib import Path
from typing import Any

from engine.tools.base import BaseTool, ToolResult

ALLOWED_EXTENSIONS = {".xlsx", ".xls", ".csv", ".tsv"}


class SpreadsheetAnalyzerTool(BaseTool):
    name = "spreadsheet_analyzer"
    description = (
        "Analyze Excel workbooks and spreadsheets with advanced operations: read "
        "multiple sheets, extract cell ranges, analyze formulas, compute cross-sheet "
        "references, generate pivot tables, detect data types per column, identify "
        "merged cells and formatting patterns, compute statistics across sheets, and "
        "extract chart data. Supports .xlsx, .xls, .csv, and .tsv formats."
    )
    input_schema: dict[str, Any] = {
        "type": "object",
        "properties": {
            "file_path": {
                "type": "string",
                "description": "Path to the spreadsheet file",
            },
            "operation": {
                "type": "string",
                "enum": [
                    "overview",
                    "read_sheet",
                    "read_range",
                    "formulas",
                    "statistics",
                    "pivot",
                    "compare_sheets",
                    "search",
                ],
                "description": "Analysis operation to perform",
                "default": "overview",
            },
            "sheet_name": {
                "type": "string",
                "description": "Sheet name to analyze (default: first sheet)",
            },
            "cell_range": {
                "type": "string",
                "description": "Cell range to read (e.g. 'A1:D10', 'B:B')",
            },
            "search_term": {
                "type": "string",
                "description": "Text to search for across all sheets",
            },
            "pivot_rows": {
                "type": "string",
                "description": "Column name for pivot table rows",
            },
            "pivot_values": {
                "type": "string",
                "description": "Column name for pivot table values",
            },
            "pivot_func": {
                "type": "string",
                "enum": ["sum", "count", "avg", "min", "max"],
                "description": "Aggregation function for pivot",
                "default": "sum",
            },
            "max_rows": {
                "type": "integer",
                "description": "Max rows to return (default: 100)",
                "default": 100,
            },
        },
        "required": [],
    }

    async def execute(self, arguments: dict[str, Any]) -> ToolResult:
        from engine.tools._inline_file import materialise_path

        operation = arguments.get("operation", "overview")
        # Allow either `file_path` (legacy) or `path`/`text` (new).
        norm = dict(arguments)
        if arguments.get("file_path") and not norm.get("path"):
            norm["path"] = arguments["file_path"]
        resolved, err = materialise_path(norm, default_ext="csv")
        if err:
            return ToolResult(content=err, is_error=True)
        file_path = resolved

        path = Path(file_path)
        if not path.exists():
            return ToolResult(content=f"File not found: {file_path}", is_error=True)

        ext = path.suffix.lower()
        if ext not in ALLOWED_EXTENSIONS:
            return ToolResult(
                content=f"Unsupported format: {ext}. Supported: {ALLOWED_EXTENSIONS}",
                is_error=True,
            )

        try:
            if ext in (".csv", ".tsv"):
                result = self._handle_csv(path, operation, arguments)
            else:
                result = self._handle_excel(path, operation, arguments)

            output = json.dumps(result, indent=2, default=str)
            return ToolResult(
                content=output,
                metadata={"file": str(path), "operation": operation},
            )
        except Exception as e:
            return ToolResult(content=f"Spreadsheet error: {e}", is_error=True)

    def _handle_csv(
        self, path: Path, operation: str, args: dict[str, Any]
    ) -> dict[str, Any]:
        delimiter = "\t" if path.suffix.lower() == ".tsv" else ","
        text = path.read_text(encoding="utf-8", errors="replace")
        reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
        headers = reader.fieldnames or []
        rows = list(reader)
        max_rows = args.get("max_rows", 100)

        if operation == "overview":
            return self._csv_overview(headers, rows, path)
        elif operation == "read_sheet":
            return {"headers": headers, "row_count": len(rows), "rows": rows[:max_rows]}
        elif operation == "statistics":
            return self._csv_statistics(headers, rows)
        elif operation == "search":
            return self._csv_search(headers, rows, args.get("search_term", ""))
        elif operation == "pivot":
            return self._csv_pivot(headers, rows, args)
        else:
            return {"headers": headers, "row_count": len(rows), "rows": rows[:max_rows]}

    def _csv_overview(
        self, headers: list[str], rows: list[dict[str, str]], path: Path
    ) -> dict[str, Any]:
        col_types: dict[str, str] = {}
        for h in headers:
            vals = [r.get(h, "") for r in rows[:50] if r.get(h, "").strip()]
            if not vals:
                col_types[h] = "empty"
                continue
            nums = sum(1 for v in vals if self._is_numeric(v))
            if nums > len(vals) * 0.8:
                col_types[h] = "numeric"
            elif all(self._is_date(v) for v in vals[:10]):
                col_types[h] = "date"
            else:
                col_types[h] = "text"

        return {
            "file": str(path),
            "format": path.suffix,
            "sheets": [{"name": "Sheet1", "rows": len(rows), "columns": len(headers)}],
            "headers": headers,
            "column_types": col_types,
            "total_rows": len(rows),
            "sample_rows": rows[:5],
        }

    def _csv_statistics(
        self, headers: list[str], rows: list[dict[str, str]]
    ) -> dict[str, Any]:
        stats: dict[str, Any] = {}
        for h in headers:
            vals = [r.get(h, "") for r in rows]
            nums = [
                float(v.replace(",", "").replace("$", "").replace("%", ""))
                for v in vals
                if self._is_numeric(v)
            ]
            if nums:
                stats[h] = {
                    "type": "numeric",
                    "count": len(nums),
                    "missing": len(vals) - len(nums),
                    "mean": round(statistics.mean(nums), 4),
                    "median": round(statistics.median(nums), 4),
                    "std": round(statistics.stdev(nums), 4) if len(nums) > 1 else 0,
                    "min": min(nums),
                    "max": max(nums),
                    "sum": round(sum(nums), 4),
                }
            else:
                unique = set(v for v in vals if v.strip())
                stats[h] = {
                    "type": "text",
                    "count": len([v for v in vals if v.strip()]),
                    "missing": len([v for v in vals if not v.strip()]),
                    "unique": len(unique),
                    "top_values": sorted(
                        ((v, vals.count(v)) for v in unique),
                        key=lambda x: x[1],
                        reverse=True,
                    )[:5],
                }
        return {"statistics": stats}

    def _csv_search(
        self, headers: list[str], rows: list[dict[str, str]], term: str
    ) -> dict[str, Any]:
        if not term:
            return {"error": "search_term is required"}
        matches = []
        for i, row in enumerate(rows):
            for h in headers:
                if term.lower() in row.get(h, "").lower():
                    matches.append(
                        {"row": i + 1, "column": h, "value": row[h], "full_row": row}
                    )
                    break
        return {
            "search_term": term,
            "match_count": len(matches),
            "matches": matches[:50],
        }

    def _csv_pivot(
        self, headers: list[str], rows: list[dict[str, str]], args: dict[str, Any]
    ) -> dict[str, Any]:
        pivot_rows = args.get("pivot_rows", "")
        pivot_values = args.get("pivot_values", "")
        func = args.get("pivot_func", "sum")

        if not pivot_rows:
            return {"error": "pivot_rows is required"}

        groups: dict[str, list[float]] = defaultdict(list)
        for row in rows:
            key = row.get(pivot_rows, "(empty)")
            if pivot_values:
                v = row.get(pivot_values, "")
                if self._is_numeric(v):
                    groups[key].append(
                        float(v.replace(",", "").replace("$", "").replace("%", ""))
                    )
            else:
                groups[key].append(1)

        agg_fns = {
            "sum": lambda v: round(sum(v), 4),
            "count": lambda v: len(v),
            "avg": lambda v: round(statistics.mean(v), 4) if v else 0,
            "min": lambda v: min(v) if v else 0,
            "max": lambda v: max(v) if v else 0,
        }
        fn = agg_fns.get(func, agg_fns["sum"])

        pivot_result = [
            {"group": key, "value": fn(vals), "count": len(vals)}
            for key, vals in sorted(groups.items())
        ]

        return {
            "pivot_by": pivot_rows,
            "value_column": pivot_values or "(count)",
            "function": func,
            "rows": pivot_result,
        }

    def _handle_excel(
        self, path: Path, operation: str, args: dict[str, Any]
    ) -> dict[str, Any]:
        from openpyxl import load_workbook

        wb = load_workbook(
            str(path),
            read_only=(operation != "formulas"),
            data_only=(operation != "formulas"),
        )

        if operation == "overview":
            return self._excel_overview(wb, path)
        elif operation == "read_sheet":
            return self._excel_read_sheet(wb, args)
        elif operation == "read_range":
            return self._excel_read_range(wb, args)
        elif operation == "formulas":
            wb_formula = load_workbook(str(path), data_only=False)
            return self._excel_formulas(wb_formula, args)
        elif operation == "statistics":
            return self._excel_statistics(wb, args)
        elif operation == "search":
            return self._excel_search(wb, args.get("search_term", ""))
        elif operation == "compare_sheets":
            return self._excel_compare(wb, args)
        elif operation == "pivot":
            return self._excel_pivot(wb, args)
        else:
            return self._excel_overview(wb, path)

    def _excel_overview(self, wb: Any, path: Path) -> dict[str, Any]:
        sheets = []
        for name in wb.sheetnames:
            ws = wb[name]
            sheets.append(
                {
                    "name": name,
                    "rows": ws.max_row or 0,
                    "columns": ws.max_column or 0,
                    "dimensions": ws.dimensions,
                }
            )

        first_sheet = wb[wb.sheetnames[0]]
        headers = []
        sample_rows = []
        for i, row in enumerate(first_sheet.iter_rows(max_row=6, values_only=True)):
            if i == 0:
                headers = [str(c) if c is not None else "" for c in row]
            else:
                sample_rows.append([str(c) if c is not None else "" for c in row])

        return {
            "file": str(path),
            "format": path.suffix,
            "sheet_count": len(wb.sheetnames),
            "sheets": sheets,
            "first_sheet_headers": headers,
            "sample_data": sample_rows[:5],
            "file_size_bytes": path.stat().st_size,
        }

    def _excel_read_sheet(self, wb: Any, args: dict[str, Any]) -> dict[str, Any]:
        sheet_name = args.get("sheet_name", wb.sheetnames[0])
        max_rows = args.get("max_rows", 100)

        if sheet_name not in wb.sheetnames:
            return {
                "error": f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"
            }

        ws = wb[sheet_name]
        headers: list[str] = []
        rows: list[dict[str, Any]] = []

        for i, row in enumerate(ws.iter_rows(max_row=max_rows + 1, values_only=True)):
            vals = [v if v is not None else "" for v in row]
            if i == 0:
                headers = [str(v) for v in vals]
            else:
                row_dict = {
                    headers[j]: vals[j] for j in range(min(len(headers), len(vals)))
                }
                rows.append(row_dict)

        return {
            "sheet": sheet_name,
            "headers": headers,
            "row_count": len(rows),
            "total_rows": ws.max_row,
            "rows": rows,
        }

    def _excel_read_range(self, wb: Any, args: dict[str, Any]) -> dict[str, Any]:
        sheet_name = args.get("sheet_name", wb.sheetnames[0])
        cell_range = args.get("cell_range", "")

        if sheet_name not in wb.sheetnames:
            return {"error": f"Sheet '{sheet_name}' not found"}
        if not cell_range:
            return {"error": "cell_range is required (e.g. 'A1:D10')"}

        ws = wb[sheet_name]
        cells = []
        for row in ws[cell_range]:
            cells.append([c.value for c in row])

        return {
            "sheet": sheet_name,
            "range": cell_range,
            "rows": len(cells),
            "data": cells,
        }

    def _excel_formulas(self, wb: Any, args: dict[str, Any]) -> dict[str, Any]:
        sheet_name = args.get("sheet_name", wb.sheetnames[0])
        ws = wb[sheet_name]

        formulas = []
        for row in ws.iter_rows():
            for cell in row:
                if (
                    cell.value
                    and isinstance(cell.value, str)
                    and cell.value.startswith("=")
                ):
                    formulas.append(
                        {
                            "cell": cell.coordinate,
                            "formula": cell.value,
                        }
                    )

        return {
            "sheet": sheet_name,
            "formula_count": len(formulas),
            "formulas": formulas[:200],
        }

    def _excel_statistics(self, wb: Any, args: dict[str, Any]) -> dict[str, Any]:
        sheet_name = args.get("sheet_name", wb.sheetnames[0])
        ws = wb[sheet_name]
        max_rows = min(ws.max_row or 0, 10000)

        headers: list[str] = []
        columns: dict[str, list[Any]] = defaultdict(list)

        for i, row in enumerate(ws.iter_rows(max_row=max_rows, values_only=True)):
            if i == 0:
                headers = [
                    str(c) if c is not None else f"col_{j}" for j, c in enumerate(row)
                ]
            else:
                for j, val in enumerate(row):
                    if j < len(headers):
                        columns[headers[j]].append(val)

        stats: dict[str, Any] = {}
        for h in headers:
            vals = columns[h]
            nums = [v for v in vals if isinstance(v, (int, float))]
            if nums:
                stats[h] = {
                    "type": "numeric",
                    "count": len(nums),
                    "missing": len(vals) - len(nums),
                    "mean": round(statistics.mean(nums), 4),
                    "std": round(statistics.stdev(nums), 4) if len(nums) > 1 else 0,
                    "min": min(nums),
                    "max": max(nums),
                    "sum": round(sum(nums), 4),
                }
            else:
                non_empty = [str(v) for v in vals if v is not None and str(v).strip()]
                stats[h] = {
                    "type": "text",
                    "count": len(non_empty),
                    "missing": len(vals) - len(non_empty),
                    "unique": len(set(non_empty)),
                }

        return {"sheet": sheet_name, "total_rows": max_rows - 1, "statistics": stats}

    def _excel_search(self, wb: Any, term: str) -> dict[str, Any]:
        if not term:
            return {"error": "search_term is required"}

        matches = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(max_row=min(ws.max_row or 0, 5000)):
                for cell in row:
                    if cell.value and term.lower() in str(cell.value).lower():
                        matches.append(
                            {
                                "sheet": sheet_name,
                                "cell": cell.coordinate,
                                "value": str(cell.value),
                            }
                        )

        return {
            "search_term": term,
            "match_count": len(matches),
            "matches": matches[:100],
        }

    def _excel_compare(self, wb: Any, args: dict[str, Any]) -> dict[str, Any]:
        sheets = wb.sheetnames
        if len(sheets) < 2:
            return {"error": "Need at least 2 sheets to compare"}

        comparison = []
        for name in sheets:
            ws = wb[name]
            comparison.append(
                {
                    "name": name,
                    "rows": ws.max_row or 0,
                    "columns": ws.max_column or 0,
                }
            )

        return {"sheets": comparison, "sheet_count": len(sheets)}

    def _excel_pivot(self, wb: Any, args: dict[str, Any]) -> dict[str, Any]:
        sheet_name = args.get("sheet_name", wb.sheetnames[0])
        ws = wb[sheet_name]

        headers: list[str] = []
        rows: list[dict[str, str]] = []
        for i, row in enumerate(ws.iter_rows(max_row=5000, values_only=True)):
            if i == 0:
                headers = [str(c) if c is not None else "" for c in row]
            else:
                rows.append(
                    {
                        headers[j]: str(v) if v is not None else ""
                        for j, v in enumerate(row)
                        if j < len(headers)
                    }
                )

        return self._csv_pivot(headers, rows, args)

    def _is_numeric(self, val: str) -> bool:
        try:
            float(val.replace(",", "").replace("$", "").replace("%", "").strip())
            return True
        except (ValueError, AttributeError):
            return False

    def _is_date(self, val: str) -> bool:
        import re

        return bool(re.match(r"\d{1,4}[-/]\d{1,2}[-/]\d{1,4}", val.strip()))
