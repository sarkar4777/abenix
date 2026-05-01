"""Export and deliver data to various destinations - email, files, S3, webhooks, databases."""

from __future__ import annotations

import csv
import io
import json
import os
from pathlib import Path
from typing import Any

from engine.tools.base import BaseTool, ToolResult

SMTP_HOST = os.environ.get("SMTP_HOST", "")
SMTP_PORT = int(os.environ.get("SMTP_PORT", "587"))
SMTP_USER = os.environ.get("SMTP_USER", "")
SMTP_PASS = os.environ.get("SMTP_PASS", "")
SMTP_FROM = os.environ.get("SMTP_FROM", "agent@abenix.dev")

AWS_ACCESS_KEY = os.environ.get("AWS_ACCESS_KEY_ID", "")
AWS_SECRET_KEY = os.environ.get("AWS_SECRET_ACCESS_KEY", "")
AWS_REGION = os.environ.get("AWS_REGION", "us-east-1")

EXPORT_DIR = os.environ.get("EXPORT_DIR", "/tmp/abenix_exports")


class DataExporterTool(BaseTool):
    name = "data_exporter"
    description = (
        "Export and deliver data to various destinations: save as file (JSON, CSV, TXT, "
        "Markdown, HTML, XLSX Excel, PDF report), send via email with attachments, upload "
        "to S3, push to webhooks, or write to databases. Supports binary formats like "
        "Excel (.xlsx) and PDF natively. Useful for delivering agent analysis results, "
        "reports, and processed data to external systems."
    )
    input_schema: dict[str, Any] = {
        "type": "object",
        "properties": {
            "destination": {
                "type": "string",
                "enum": ["file", "email", "s3", "webhook", "database"],
                "description": "Export destination",
            },
            "data": {
                "description": "Data to export (string, object, or array)",
            },
            "format": {
                "type": "string",
                "enum": ["json", "csv", "txt", "markdown", "html", "xlsx", "pdf"],
                "description": "Output format. Use xlsx for Excel spreadsheets, pdf for PDF reports.",
                "default": "json",
            },
            "filename": {
                "type": "string",
                "description": "Output filename (auto-generated if omitted)",
            },
            "email_to": {
                "type": "string",
                "description": "Recipient email address(es), comma-separated",
            },
            "email_subject": {
                "type": "string",
                "description": "Email subject line",
            },
            "email_body": {
                "type": "string",
                "description": "Email body text (the data will be attached)",
            },
            "s3_bucket": {
                "type": "string",
                "description": "S3 bucket name",
            },
            "s3_key": {
                "type": "string",
                "description": "S3 object key/path",
            },
            "webhook_url": {
                "type": "string",
                "description": "Webhook URL to POST data to",
            },
            "webhook_headers": {
                "type": "object",
                "description": "Additional headers for webhook",
            },
            "db_connection_string": {
                "type": "string",
                "description": "Database connection string",
            },
            "db_table": {
                "type": "string",
                "description": "Database table name",
            },
        },
        "required": ["destination", "data"],
    }

    async def execute(self, arguments: dict[str, Any]) -> ToolResult:
        destination = arguments.get("destination", "")
        data = arguments.get("data")

        if not destination:
            return ToolResult(content="Error: destination is required", is_error=True)
        if data is None:
            return ToolResult(content="Error: data is required", is_error=True)

        handlers = {
            "file": self._export_file,
            "email": self._export_email,
            "s3": self._export_s3,
            "webhook": self._export_webhook,
            "database": self._export_database,
        }

        fn = handlers.get(destination)
        if not fn:
            return ToolResult(content=f"Unknown destination: {destination}", is_error=True)

        try:
            result = await fn(data, arguments)
            output = json.dumps(result, indent=2, default=str)
            return ToolResult(content=output, metadata={"destination": destination})
        except Exception as e:
            return ToolResult(content=f"Export error: {e}", is_error=True)

    def _format_data(self, data: Any, fmt: str) -> str:
        if fmt == "json":
            if isinstance(data, str):
                try:
                    data = json.loads(data)
                except json.JSONDecodeError:
                    return data
            return json.dumps(data, indent=2, default=str)
        elif fmt == "csv":
            return self._to_csv(data)
        elif fmt == "markdown":
            return self._to_markdown(data)
        elif fmt == "html":
            return self._to_html(data)
        else:
            return str(data)

    def _to_csv(self, data: Any) -> str:
        output = io.StringIO()
        if isinstance(data, list) and data and isinstance(data[0], dict):
            headers = list(data[0].keys())
            writer = csv.DictWriter(output, fieldnames=headers)
            writer.writeheader()
            writer.writerows(data)
        elif isinstance(data, dict):
            writer = csv.writer(output)
            for key, val in data.items():
                writer.writerow([key, val])
        else:
            output.write(str(data))
        return output.getvalue()

    def _to_markdown(self, data: Any) -> str:
        if isinstance(data, list) and data and isinstance(data[0], dict):
            headers = list(data[0].keys())
            lines = ["| " + " | ".join(headers) + " |"]
            lines.append("| " + " | ".join("---" for _ in headers) + " |")
            for row in data:
                lines.append("| " + " | ".join(str(row.get(h, "")) for h in headers) + " |")
            return "\n".join(lines)
        elif isinstance(data, dict):
            lines = []
            for key, val in data.items():
                if isinstance(val, dict):
                    lines.append(f"\n## {key}\n")
                    for k2, v2 in val.items():
                        lines.append(f"- **{k2}**: {v2}")
                else:
                    lines.append(f"- **{key}**: {val}")
            return "\n".join(lines)
        return str(data)

    def _to_html(self, data: Any) -> str:
        if isinstance(data, list) and data and isinstance(data[0], dict):
            headers = list(data[0].keys())
            rows_html = ""
            for row in data:
                cells = "".join(f"<td>{row.get(h, '')}</td>" for h in headers)
                rows_html += f"<tr>{cells}</tr>"
            header_html = "".join(f"<th>{h}</th>" for h in headers)
            return f"<table><thead><tr>{header_html}</tr></thead><tbody>{rows_html}</tbody></table>"
        return f"<pre>{json.dumps(data, indent=2, default=str)}</pre>"

    def _to_xlsx(self, data: Any) -> bytes:
        """Generate an Excel workbook from data using openpyxl."""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active

        # Parse string data if needed
        if isinstance(data, str):
            try:
                data = json.loads(data)
            except json.JSONDecodeError:
                pass

        # Style definitions
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9"),
        )
        alt_fill = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")

        if isinstance(data, list) and data and isinstance(data[0], dict):
            # Array of objects → table with headers
            ws.title = "Data"
            headers = list(data[0].keys())
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=str(header).replace("_", " ").title())
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = thin_border
            for row_idx, row in enumerate(data, 2):
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=row.get(header, ""))
                    cell.border = thin_border
                    if row_idx % 2 == 0:
                        cell.fill = alt_fill
            # Auto-fit column widths
            for col_idx, header in enumerate(headers, 1):
                max_len = max(len(str(header)), *(len(str(row.get(header, ""))) for row in data)) + 2
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len, 50)
        elif isinstance(data, dict):
            # Dict → key-value pairs, or multiple sheets for nested dicts
            ws.title = "Summary"
            ws.cell(row=1, column=1, value="Field").font = header_font
            ws.cell(row=1, column=1).fill = header_fill
            ws.cell(row=1, column=1).border = thin_border
            ws.cell(row=1, column=2, value="Value").font = header_font
            ws.cell(row=1, column=2).fill = header_fill
            ws.cell(row=1, column=2).border = thin_border
            row_num = 2
            for key, val in data.items():
                if isinstance(val, (list, dict)):
                    # Nested structures get their own sheet
                    if isinstance(val, list) and val and isinstance(val[0], dict):
                        detail_ws = wb.create_sheet(title=str(key)[:31])
                        detail_headers = list(val[0].keys())
                        for ci, h in enumerate(detail_headers, 1):
                            c = detail_ws.cell(row=1, column=ci, value=str(h).replace("_", " ").title())
                            c.font = header_font
                            c.fill = header_fill
                            c.border = thin_border
                        for ri, r in enumerate(val, 2):
                            for ci, h in enumerate(detail_headers, 1):
                                c = detail_ws.cell(row=ri, column=ci, value=r.get(h, ""))
                                c.border = thin_border
                    cell = ws.cell(row=row_num, column=1, value=str(key))
                    cell.border = thin_border
                    ref_cell = ws.cell(row=row_num, column=2, value=f"See '{key}' sheet" if isinstance(val, list) else json.dumps(val, default=str)[:500])
                    ref_cell.border = thin_border
                else:
                    cell = ws.cell(row=row_num, column=1, value=str(key))
                    cell.border = thin_border
                    val_cell = ws.cell(row=row_num, column=2, value=val)
                    val_cell.border = thin_border
                if row_num % 2 == 0:
                    ws.cell(row=row_num, column=1).fill = alt_fill
                    ws.cell(row=row_num, column=2).fill = alt_fill
                row_num += 1
            ws.column_dimensions["A"].width = 30
            ws.column_dimensions["B"].width = 60
        else:
            ws.title = "Data"
            ws.cell(row=1, column=1, value=str(data))

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def _to_pdf(self, data: Any) -> bytes:
        """Generate a PDF report from data. Uses basic text layout without external deps."""
        # Parse string data if needed
        if isinstance(data, str):
            try:
                data = json.loads(data)
            except json.JSONDecodeError:
                pass

        # Simple PDF generation using minimal PDF spec (no external deps)
        lines: list[str] = []
        lines.append("Abenix Export Report")
        lines.append("=" * 50)
        lines.append("")

        if isinstance(data, list) and data and isinstance(data[0], dict):
            headers = list(data[0].keys())
            lines.append("  |  ".join(h.replace("_", " ").title() for h in headers))
            lines.append("-" * 80)
            for row in data:
                lines.append("  |  ".join(str(row.get(h, ""))[:30] for h in headers))
        elif isinstance(data, dict):
            for key, val in data.items():
                if isinstance(val, (dict, list)):
                    lines.append(f"\n--- {key} ---")
                    lines.append(json.dumps(val, indent=2, default=str)[:2000])
                else:
                    lines.append(f"{key}: {val}")
        else:
            lines.append(str(data))

        text = "\n".join(lines)

        # Minimal valid PDF with text content
        text_escaped = text.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")
        # Split into pages of ~60 lines each
        all_lines = text_escaped.split("\n")
        page_size = 60
        pages = [all_lines[i:i + page_size] for i in range(0, len(all_lines), page_size)]
        if not pages:
            pages = [[""]]

        objects: list[str] = []
        # Obj 1: Catalog
        objects.append("1 0 obj\n<< /Type /Catalog /Pages 2 0 R >>\nendobj")
        # Obj 2: Pages (placeholder, filled later)
        page_refs = " ".join(f"{3 + i * 2} 0 R" for i in range(len(pages)))
        objects.append(f"2 0 obj\n<< /Type /Pages /Kids [{page_refs}] /Count {len(pages)} >>\nendobj")
        # Obj for each page + content
        obj_num = 3
        font_obj = obj_num + len(pages) * 2
        for page_lines in pages:
            content_obj = obj_num + 1
            y = 750
            stream_lines = [f"BT /F1 10 Tf"]
            for line in page_lines:
                stream_lines.append(f"1 0 0 1 50 {y} Tm ({line}) Tj")
                y -= 14
            stream_lines.append("ET")
            stream = "\n".join(stream_lines)
            objects.append(f"{obj_num} 0 obj\n<< /Type /Page /Parent 2 0 R /MediaBox [0 0 612 792] /Contents {content_obj} 0 R /Resources << /Font << /F1 {font_obj} 0 R >> >> >>\nendobj")
            objects.append(f"{content_obj} 0 obj\n<< /Length {len(stream)} >>\nstream\n{stream}\nendstream\nendobj")
            obj_num += 2
        # Font object
        objects.append(f"{font_obj} 0 obj\n<< /Type /Font /Subtype /Type1 /BaseFont /Courier >>\nendobj")

        body = "\n".join(objects)
        xref_offset = len(f"%PDF-1.4\n{body}\n")
        num_objs = font_obj + 1
        xref = f"xref\n0 {num_objs}\n0000000000 65535 f \n"
        # Simplified xref (viewers are tolerant)
        for i in range(1, num_objs):
            xref += f"{str(i * 100).zfill(10)} 00000 n \n"
        trailer = f"trailer\n<< /Size {num_objs} /Root 1 0 R >>\nstartxref\n{xref_offset}\n%%EOF"
        pdf_content = f"%PDF-1.4\n{body}\n{xref}{trailer}"
        return pdf_content.encode("latin-1")

    async def _export_file(self, data: Any, args: dict[str, Any]) -> dict[str, Any]:
        fmt = args.get("format", "json")
        filename = args.get("filename", "")

        if not filename:
            import uuid as _uuid
            ext = {"json": ".json", "csv": ".csv", "txt": ".txt", "markdown": ".md",
                   "html": ".html", "xlsx": ".xlsx", "pdf": ".pdf"}
            filename = f"export_{_uuid.uuid4().hex[:8]}{ext.get(fmt, '.txt')}"

        # Generate file content
        if fmt == "xlsx":
            content_bytes = self._to_xlsx(data)
        elif fmt == "pdf":
            content_bytes = self._to_pdf(data)
        else:
            content_text = self._format_data(data, fmt)
            content_bytes = content_text.encode("utf-8")

        # Write to local export dir (for pipeline node access within same execution)
        os.makedirs(EXPORT_DIR, exist_ok=True)
        filepath = os.path.join(EXPORT_DIR, filename)
        with open(filepath, "wb") as f:
            f.write(content_bytes)

        # Persist to storage backend (S3/Azure for K8s, local for dev)
        storage_uri = filepath
        download_url = f"/api/files/export/{filename}"
        try:
            from engine.storage import get_storage
            storage = get_storage()
            if storage.backend != "local":
                content_type = {
                    "json": "application/json", "csv": "text/csv", "txt": "text/plain",
                    "markdown": "text/markdown", "html": "text/html",
                    "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    "pdf": "application/pdf",
                }.get(fmt, "application/octet-stream")
                storage_uri = await storage.upload(
                    tenant_id="exports",
                    path=f"exports/{filename}",
                    data=content_bytes,
                    content_type=content_type,
                )
                download_url = await storage.get_download_url(storage_uri, expires=86400)
        except Exception:
            pass  # Storage service unavailable — local file still works

        return {
            "status": "success",
            "destination": "file",
            "file_path": filepath,
            "storage_uri": storage_uri,
            "download_url": download_url,
            "filename": filename,
            "format": fmt,
            "size_bytes": len(content_bytes),
        }

    async def _export_email(self, data: Any, args: dict[str, Any]) -> dict[str, Any]:
        email_to = args.get("email_to", "")
        subject = args.get("email_subject", "Abenix Export")
        body = args.get("email_body", "Please find the attached export data.")
        fmt = args.get("format", "json")

        if not email_to:
            return {"error": "email_to is required"}

        if not SMTP_HOST:
            file_result = await self._export_file(data, args)
            return {
                "status": "mock",
                "message": "Email not configured (SMTP_HOST not set). Data saved to file instead.",
                "would_send_to": email_to,
                "subject": subject,
                "file_saved": file_result.get("file_path"),
            }

        import smtplib
        from email.mime.multipart import MIMEMultipart
        from email.mime.text import MIMEText
        from email.mime.base import MIMEBase
        from email import encoders

        msg = MIMEMultipart()
        msg["From"] = SMTP_FROM
        msg["To"] = email_to
        msg["Subject"] = subject
        msg.attach(MIMEText(body, "plain"))

        content = self._format_data(data, fmt)
        ext = {"json": ".json", "csv": ".csv", "txt": ".txt", "markdown": ".md", "html": ".html"}
        filename = args.get("filename", f"export{ext.get(fmt, '.txt')}")

        attachment = MIMEBase("application", "octet-stream")
        attachment.set_payload(content.encode("utf-8"))
        encoders.encode_base64(attachment)
        attachment.add_header("Content-Disposition", f"attachment; filename={filename}")
        msg.attach(attachment)

        with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as server:
            server.starttls()
            if SMTP_USER and SMTP_PASS:
                server.login(SMTP_USER, SMTP_PASS)
            server.sendmail(SMTP_FROM, email_to.split(","), msg.as_string())

        return {
            "status": "success",
            "destination": "email",
            "sent_to": email_to,
            "subject": subject,
            "attachment": filename,
        }

    async def _export_s3(self, data: Any, args: dict[str, Any]) -> dict[str, Any]:
        bucket = args.get("s3_bucket", "")
        key = args.get("s3_key", "")
        fmt = args.get("format", "json")

        if not bucket or not key:
            return {"error": "s3_bucket and s3_key are required"}

        if not AWS_ACCESS_KEY:
            file_result = await self._export_file(data, args)
            return {
                "status": "mock",
                "message": "AWS credentials not configured. Data saved to file instead.",
                "would_upload_to": f"s3://{bucket}/{key}",
                "file_saved": file_result.get("file_path"),
            }

        import boto3
        content = self._format_data(data, fmt)
        s3 = boto3.client("s3", region_name=AWS_REGION)
        s3.put_object(Bucket=bucket, Key=key, Body=content.encode("utf-8"))

        return {
            "status": "success",
            "destination": "s3",
            "bucket": bucket,
            "key": key,
            "size_bytes": len(content.encode("utf-8")),
            "url": f"s3://{bucket}/{key}",
        }

    async def _export_webhook(self, data: Any, args: dict[str, Any]) -> dict[str, Any]:
        webhook_url = args.get("webhook_url", "")
        extra_headers = args.get("webhook_headers", {})

        if not webhook_url:
            return {"error": "webhook_url is required"}

        import aiohttp

        headers = {
            "Content-Type": "application/json",
            "User-Agent": "Abenix-Export/1.0",
            **extra_headers,
        }

        payload = data if isinstance(data, (dict, list)) else {"data": data}

        async with aiohttp.ClientSession() as session:
            async with session.post(
                webhook_url,
                json=payload,
                headers=headers,
                timeout=aiohttp.ClientTimeout(total=15),
            ) as resp:
                return {
                    "status": "success",
                    "destination": "webhook",
                    "url": webhook_url,
                    "response_status": resp.status,
                    "response_body": await resp.text()[:1000],
                }

    async def _export_database(self, data: Any, args: dict[str, Any]) -> dict[str, Any]:
        conn_str = args.get("db_connection_string", "")
        table = args.get("db_table", "")

        if not conn_str or not table:
            return {"error": "db_connection_string and db_table are required"}

        if not isinstance(data, list) or not data or not isinstance(data[0], dict):
            return {"error": "Data must be an array of objects for database export"}

        file_result = await self._export_file(data, {**args, "format": "json"})
        return {
            "status": "pending",
            "message": "Database export prepared. Data saved to file for import.",
            "table": table,
            "row_count": len(data),
            "columns": list(data[0].keys()),
            "file_saved": file_result.get("file_path"),
        }
